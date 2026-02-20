nombre = input("ingrese su nombre completo: ")
matricula = input("ingrese su matricula: ")
carrera = input("ingrese su carrera: ")
semestre = input("ingrese su semestre de ingreso: ")
fecha = input("ingrese la fecha de hoy: ")

print("\n--- datos registrados ---")
print("nombre:", nombre)
print("matricula:", matricula)
print("carrera:", carrera)
print("semestre:", semestre)
print("fecha:", fecha)

from openpyxl import Workbook, load_workbook
import os

archivo = "registro_alumnos.xlsx"

# crear excel si no existe
if not os.path.exists(archivo):
    libro = Workbook()
    hoja = libro.active
    hoja.append(["Nombre", "Matricula", "Carrera", "semestre", "Fecha"])
    libro.save(archivo)


# abrir excel y guardar datos
libro = load_workbook(archivo)
hoja = libro.active

hoja.append([nombre, matricula, carrera, semestre, fecha])
libro.save(archivo)


print("\nDATOS GUARDADOS EN EXCEL:\n")

for fila in hoja.iter_rows(values_only=True):
    print(fila)

